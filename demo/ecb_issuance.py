"""Demo pipeline: gross issuance of debt securities, from the ECB's public API.

This is a deliberately small but complete instance of the architecture the
library exists to serve: fetch -> transform -> validate -> formatted Excel,
with every sentinel check wired in where it earns its keep.

    python demo/ecb_issuance.py                    # live API
    python demo/ecb_issuance.py --offline          # cached fixture (CI uses this)
    python demo/ecb_issuance.py --sabotage units   # watch the sentinel catch it

``--sabotage`` injects a classic silent error into the extraction path
(``units``: a thousand-fold scaling slip; ``fx``: a ~1.3% multiplicative
convention error; ``drop``: one country quietly lost). None of them crash
anything — which is the point. Run one and read the benchmark report.

Data: ECB Securities Issues Statistics (CSEC), monthly gross issuance of
debt securities, total economy, eight euro-area countries. Publicly
available, no key required.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import sys
from pathlib import Path
from typing import Mapping

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from sentinel import flag_table, run_benchmark

HERE = Path(__file__).parent
FIXTURE = HERE / "fixtures" / "csec_gross_s1.csv"
REFERENCE = HERE / "fixtures" / "reference_2024.json"
OUTPUT = HERE / "output" / "gross_issuance.xlsx"

COUNTRIES = {
    "AT": "Austria",
    "BE": "Belgium",
    "DE": "Germany",
    "ES": "Spain",
    "FR": "France",
    "IT": "Italy",
    "NL": "Netherlands",
    "PT": "Portugal",
}

# One dot-separated SDMX key defines a series across 18 dimensions. The two
# fields that matter here: ACCOUNTING_ENTRY = LI (gross issuance) and
# CURRENCY_DENOM, where we pull EUR (euro-denominated), X1 (foreign-currency)
# and _T (total) separately — because _T should equal EUR + X1, and an
# identity the source itself publishes is a free cross-check (see below).
KEY = "M.N.{areas}.W0.S1.S1.N.LI.F.F3.T._Z.EUR.{denoms}.F.V.N._T"
API = "https://data-api.ecb.europa.eu/service/data/CSEC/" + KEY


def fetch(offline: bool) -> list[dict[str, str]]:
    if offline:
        text = FIXTURE.read_text()
    else:
        url = API.format(areas="+".join(COUNTRIES), denoms="EUR+X1+_T")
        response = requests.get(
            url, params={"format": "csvdata", "startPeriod": "2014-01"}, timeout=90
        )
        response.raise_for_status()
        text = response.text
    return list(csv.DictReader(io.StringIO(text)))


def annual_totals(records: list[dict[str, str]], denom: str) -> dict[str, dict[int, float]]:
    """Monthly EUR mn -> calendar-year EUR bn, per country.

    Years with fewer than 12 monthly observations are dropped rather than
    summed: an annual figure built from eleven months is the archetypal
    plausible-but-wrong number — smaller than the truth, alarming to no one.
    """
    monthly: dict[str, dict[int, list[float]]] = {name: {} for name in COUNTRIES.values()}
    for r in records:
        if r["CURRENCY_DENOM"] != denom or r["REF_AREA"] not in COUNTRIES:
            continue
        year = int(r["TIME_PERIOD"][:4])
        monthly[COUNTRIES[r["REF_AREA"]]].setdefault(year, []).append(float(r["OBS_VALUE"]))
    return {
        country: {
            year: round(sum(values) / 1000, 3)
            for year, values in sorted(years.items())
            if len(values) == 12
        }
        for country, years in monthly.items()
    }


def check_identity(records: list[dict[str, str]]) -> list[str]:
    """Verify EUR + X1 = _T per country-year, an identity the source publishes.

    A benchmark diff can only catch what differs between two runs; an error
    inherited by both is invisible to it. Identities internal to the source
    are a different *kind* of check, so their blind spots don't overlap.
    Published values carry 3 decimals, so a summed total may differ by
    rounding — the tolerance covers that, nothing more.
    """
    eur = annual_totals(records, "EUR")
    x1 = annual_totals(records, "X1")
    total = annual_totals(records, "_T")
    problems = []
    for country in COUNTRIES.values():
        for year, t in total[country].items():
            if year in eur[country] and year in x1[country]:
                gap = abs(eur[country][year] + x1[country][year] - t)
                if gap > 0.002:  # two rounding steps of headroom at 3 decimals
                    problems.append(
                        f"{country} {year}: EUR + X1 differs from _T by {gap:.3f} EUR bn"
                    )
    return problems


def extract_year(records: list[dict[str, str]], year: int, sabotage: str | None) -> dict:
    """The single extraction path used for BOTH the benchmark and the new year.

    The sabotage hook stands in for the real-world ways this path silently
    goes wrong; because the benchmark year flows through the same code, the
    corruption is caught on known answers before the new year inherits it.
    """
    table = annual_totals(records, "_T")
    column = {country: years.get(year) for country, years in table.items()}
    if sabotage == "units":
        column = {c: (v / 1000 if v is not None else None) for c, v in column.items()}
    elif sabotage == "fx":
        column = {c: (v * 1.013 if v is not None else None) for c, v in column.items()}
    elif sabotage == "drop":
        column["Netherlands"] = None
    return column


def write_workbook(
    table: Mapping[str, Mapping[int, float]],
    new_year: int,
    new_column: Mapping[str, float | None],
    flagged: set[str],
    report_lines: list[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gross issuance (EUR bn)"
    years = sorted({y for years in table.values() for y in years})
    yellow = PatternFill("solid", start_color="FFFF00")
    bold = Font(bold=True)

    ws.append(["Country", *years, new_year])
    for cell in ws[1]:
        cell.font = bold
    for country, history in table.items():
        ws.append([country, *[history.get(y) for y in years], new_column.get(country)])
        if country in flagged:
            ws.cell(row=ws.max_row, column=len(years) + 2).fill = yellow
    ws.column_dimensions["A"].width = 16

    notes = wb.create_sheet("Run report")
    notes.append(["pipeline-sentinel run report"])
    notes["A1"].font = bold
    for line in report_lines:
        notes.append([line])
    notes.column_dimensions["A"].width = 110

    OUTPUT.parent.mkdir(exist_ok=True)
    wb.save(OUTPUT)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--offline", action="store_true", help="use the cached fixture")
    parser.add_argument("--sabotage", choices=["units", "fx", "drop"], default=None)
    args = parser.parse_args()

    records = fetch(args.offline)
    report: list[str] = []

    # 1. Different-kind check: the source's own published identity.
    identity_problems = check_identity(records)
    report.append(
        "identity EUR + X1 = _T: "
        + ("holds for every country-year" if not identity_problems else "VIOLATED")
    )
    report.extend("    " + p for p in identity_problems)

    # 2. Benchmark: reproduce the reference year through the same code path
    #    that will produce the new year. Trust is transitive along it.
    reference: dict[str, float] = json.loads(REFERENCE.read_text())
    benchmark_year = 2024
    result = run_benchmark(
        lambda period: extract_year(records, int(period), args.sabotage),
        reference,
        str(benchmark_year),
    )
    report.extend(result.lines())

    # 3. Tolerance bands: judge the new year against each row's own history.
    table = annual_totals(records, "_T")
    new_year = max(y for years in table.values() for y in years)
    history = {c: [years.get(y) for y in sorted(set().union(*map(set, table.values()))) if y < new_year]
               for c, years in table.items()}
    new_column = extract_year(records, new_year, args.sabotage)
    flags = flag_table(history, new_column)
    if flags:
        report.append(f"tolerance flags for {new_year} ({len(flags)}):")
        report.extend("    " + str(f) for f in flags)
    else:
        report.append(f"tolerance flags for {new_year}: none — every change within its row's band")

    history_only = {c: {y: v for y, v in years.items() if y < new_year} for c, years in table.items()}
    write_workbook(history_only, new_year, new_column, {f.label for f in flags}, report)

    print("\n".join(report))
    print(f"\nworkbook written to {OUTPUT}")
    return 0 if (result.clean and not identity_problems) else 1


if __name__ == "__main__":
    sys.exit(main())
